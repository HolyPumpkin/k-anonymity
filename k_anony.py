from ast import operator
from cv2 import dft
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook

# 检测数据集是否满足k-匿名
def support_k(df,k):
    for i in df:
        count = 0
        flag = 1
        for j in df:
            for c in range(5):
                if i[c] != j[c]:
                    flag = 0
                    break
            if flag:
                count += 1
        if count < k:
            return False
    return True

def generate(df,col):
    return df[col].apply(lambda x:int(int(x/(10**1))*(10**1)))

def generate2(df,col):
    return df[col].apply(lambda x:'>=90'if x>=90 else '<90')

def generate3(df,col):
    return df[col].apply(lambda x:'*')

wb = load_workbook("data.xlsx")
sheets = wb.worksheets
print(sheets)

sheet1 = sheets[0]
print(sheet1)
attri = []

all_col = []
for i in ['B','C','D','E','F']:
    col_i = []
    for col in sheet1[i]:
        col_i.append(col.value)
    attri.append(col_i.pop(0))
    all_col.append(col_i)
# print(all_col)
# print(attri)
    # 将读取到的属性与属性名映射为一个字典集
pre_data = {attri[i]:all_col[i] for i in range(5)}
# print(pre_data)

df = pd.DataFrame(pre_data,columns=['性别','语文成绩','数学成绩','外语成绩','平均成绩'])

print("原始数据如下:")
print(df)
print("原始数据满足5-匿名:")
print(support_k(df.values.tolist(),5))


df['语文成绩'] = generate(df,'语文成绩')
print("对语文成绩进行泛化后满足5-匿名:")
print(support_k(df.values.tolist(),5))

df['数学成绩'] = generate(df,'数学成绩')
print("对数学成绩进行泛化后满足5-匿名:")
print(support_k(df.values.tolist(),5))

df['外语成绩'] = generate(df,'外语成绩')
print("对外语成绩进行泛化后满足5-匿名:")
print(support_k(df.values.tolist(),5))

df['平均成绩'] = generate(df,'平均成绩')
print("对平均成绩进行泛化后满足5-匿名:")
print(support_k(df.values.tolist(),5))

print("第一轮泛化后的结果:")
print(df)

df['语文成绩'] = generate2(df,'语文成绩')
print("对语文成绩再次进行泛化后满足5-匿名:")
print(support_k(df.values.tolist(),5))

df['数学成绩'] = generate2(df,'数学成绩')
print("对数学成绩再次进行泛化后满足5-匿名:")
print(support_k(df.values.tolist(),5))

df['外语成绩'] = generate2(df,'外语成绩')
print("对外语成绩再次进行泛化后满足5-匿名:")
print(support_k(df.values.tolist(),5))

df['平均成绩'] = generate2(df,'平均成绩')
print("对平均成绩再次进行泛化后满足5-匿名:")
print(support_k(df.values.tolist(),5))

df['性别'] = generate3(df,'性别')
print("对性别进行泛化后满足5-匿名:")
print(df)

wb2 = Workbook()
ws2 = wb2.active

data_list = df.values.tolist()

ws2['A1'] = '性别'
ws2['B1'].value = '语文成绩'
ws2['C1'].value = '数学成绩'
ws2['D1'].value = '外语成绩'
ws2['E1'].value = '平均成绩'
for i in data_list:
    ws2.append(i)

wb2.save("published_data.xlsx")






